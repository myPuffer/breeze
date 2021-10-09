import os
import json

from openpyxl import load_workbook
from datetime import timezone, datetime, timedelta
from colorama import Fore, Back, Style, init as colorama_init

colorama_init()

CUR_DIR = os.path.dirname(__file__)
XLSX_DIR = ''
JSON_DIR = ''
LUAS_DIR = ''

def warn(s):
	print(Fore.YELLOW+s)


def transxlsx(file):
	path = os.path.join(XLSX_DIR, file)
	wb = load_workbook(filename = path, data_only=True)
	tbl = wb['data']
	data, pk = parse_sheet(tbl)

	output_json(data, file)

	lua = 'return ' + to_lua(data, '', pk) + '\r\n'
	output_lua(lua, file)


def parse_sheet(tbl):
	keys = []
	typs = []
	for x in range(tbl.max_column):
		key = tbl.cell(2, x+1).value
		typ = tbl.cell(3, x+1).value
		if typ is not None:
			typ = typ.lower()
		keys.append(key)
		typs.append(typ)

	data = []
	
	for r in tbl.iter_rows(min_row=6):
		if r[0].value is None: 
			continue

		d = {}
		i = 0
		for x in r:
			key = keys[i]
			typ = typs[i]
			i += 1

			if typ is None: continue
			val = x.value
			if typ == 'json':
				try:
					if val == '':
						val = None
					elif val is not None:
						val = json.loads(val)					
				except Exception as e:
					warn(f"  ! cell <${x.coordinate}> val:{val} parse as json failed:")
					warn(f"  ! - {e}")

			elif typ =='string':
				if val is None:
					val = ''
				val = str(val)

			elif typ == 'int':
				val = int(val) if val is not None else None

			elif typ == 'bool':
				val = bool(val)

			d[key] = val

		data.append(d)

	return data, keys[0]

def to_lua(data, indent, id=None):
	
	if type(data)==list:
		lines = []
		lines.append("{")
		idx = 1
		for x in data:			
			l = ''
			if id is not None:
				key = x[id]
				l = f'{indent}\t[{key}] = '
			else:
				l = f'{indent}\t[{idx}] = '
				idx += 1

			l += to_lua(x, indent+'\t')
			l += ','
			lines.append(l)
		lines.append(indent+'}')
		return '\r\n'.join(lines)

	if type(data)==dict:
		lines = []
		lines.append('{')
		for kv in data.items():
			key = kv[0]
			val = kv[1]
			l = f'{indent}\t{key} = '
			l += to_lua(val, indent+'\t')
			l += ','
			lines.append(l)

		lines.append(indent+'}')
		return '\r\n'.join(lines)

	elif type(data)==float:
		return str(data)
	elif type(data)==int:
		return str(data)
	elif type(data)==str:
		return json.dumps(data, ensure_ascii=False)
	elif type(data)==bool:
		return "true" if data else "false"
	elif data is None:
		return "nil"
	elif type(data)==datetime:
		tm = data.replace(tzinfo=timezone(timedelta(hours=8)))
		ts = int(tm.timestamp())
		return f'os.date("!*t", {ts})'
	else:
		print(type(data))


def output_lua(content, file):
	path = os.path.join(LUAS_DIR, file.replace('.xlsx', '.lua'))
	f = open(path, 'wb')
	f.write(content.encode('utf-8'))
	f.close()


def output_json(data, file):
	js = json.dumps(data, indent='\t', separators=(',', ':'), ensure_ascii=False, default=str)
	path = os.path.join(JSON_DIR, file.replace('.xlsx', '.json'))
	f = open(path, 'wb')
	f.write(js.encode('utf-8'))
	f.close()
	

def trans_config():
	for x in os.scandir(XLSX_DIR):
		f = x.name
		if f.endswith('.xlsx') and not f.startswith('~$'):
			print(f"{Fore.GREEN}>{Style.RESET_ALL} TRANS: {f}")
			transxlsx(f)


if __name__ == '__main__':
	import time
	start_time = time.time()
	trans_config()
	print('>> Finished in %.1fs' % (time.time()-start_time))

	print(f'{Fore.BLUE} ======== DONE ======== ')
		
